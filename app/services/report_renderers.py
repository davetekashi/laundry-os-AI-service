import os
from html import escape
from io import BytesIO
from threading import Lock
from datetime import datetime

os.environ.setdefault("MPLCONFIGDIR", "/tmp/laundry-os-matplotlib")
os.environ.setdefault("XDG_CACHE_HOME", "/tmp/laundry-os-cache")

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.services.generated_report_data import ChartSpec, ReportDataset, ReportNarrative


NAVY = colors.HexColor("#123047")
BLUE = colors.HexColor("#159DD8")
PALE_BLUE = colors.HexColor("#EAF6FB")
MUTED = colors.HexColor("#5B6B75")
GRID = colors.HexColor("#D9E3E8")
CHART_RENDER_LOCK = Lock()


def _render_chart(spec: ChartSpec) -> BytesIO:
    figure, axis = plt.subplots(figsize=(7.2, 3.2))
    palette = ["#159DD8", "#33B98A", "#F0A03C", "#E35D6A", "#64748B", "#7CB9D1", "#B7C85B"]

    if spec.kind == "pie":
        axis.pie(
            spec.values,
            labels=spec.labels,
            autopct="%1.0f%%",
            startangle=90,
            colors=palette[: len(spec.values)],
            textprops={"fontsize": 8},
        )
        axis.axis("equal")
    elif spec.kind == "bar":
        axis.bar(spec.labels, spec.values, color=palette[0])
        axis.tick_params(axis="x", labelrotation=25, labelsize=8)
        axis.grid(axis="y", alpha=0.2)
    else:
        axis.plot(spec.labels, spec.values, color=palette[0], marker="o", linewidth=2)
        axis.fill_between(spec.labels, spec.values, color=palette[0], alpha=0.12)
        axis.tick_params(axis="x", labelrotation=25, labelsize=8)
        axis.grid(axis="y", alpha=0.2)

    axis.set_title(spec.title, loc="left", fontsize=12, fontweight="bold", color="#123047")
    axis.spines[["top", "right"]].set_visible(False)
    figure.tight_layout()
    output = BytesIO()
    figure.savefig(output, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(figure)
    output.seek(0)
    return output


def render_chart(spec: ChartSpec) -> BytesIO:
    # pyplot uses process-global state, so serialize chart rendering across requests.
    with CHART_RENDER_LOCK:
        return _render_chart(spec)


def render_pdf_report(
    dataset: ReportDataset,
    laundry_name: str,
    period_label: str,
    narrative: ReportNarrative,
) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=13 * mm,
        bottomMargin=13 * mm,
        title=dataset.title,
        author="Laundry OS",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=22, leading=26, textColor=NAVY, alignment=TA_CENTER,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"], fontSize=10,
        leading=14, textColor=MUTED, alignment=TA_CENTER,
    )
    heading_style = ParagraphStyle(
        "ReportHeading", parent=styles["Heading2"], fontSize=13,
        leading=16, textColor=NAVY, spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "ReportBody", parent=styles["BodyText"], fontSize=9.5,
        leading=14, textColor=colors.HexColor("#24343D"),
    )
    insight_style = ParagraphStyle(
        "ChartInsight", parent=body_style, fontSize=9.2, leading=13,
        textColor=colors.HexColor("#24343D"), leftIndent=3 * mm,
        rightIndent=3 * mm, borderColor=GRID, borderWidth=0.5,
        borderPadding=7, backColor=colors.HexColor("#F7FAFB"),
    )
    finding_style = ParagraphStyle(
        "Finding", parent=body_style, leftIndent=5 * mm, firstLineIndent=-3 * mm,
        spaceAfter=4,
    )

    story = [
        Paragraph(escape(dataset.title), title_style),
        Paragraph(escape(f"{laundry_name} | {period_label}"), subtitle_style),
        Spacer(1, 7 * mm),
    ]

    metric_cells = []
    for label, value in dataset.metrics:
        metric_cells.append(Paragraph(
            f'<font size="8" color="#5B6B75">{escape(label)}</font><br/>'
            f'<font size="15" color="#123047"><b>{escape(value)}</b></font>',
            ParagraphStyle("Metric", parent=body_style, leading=18, alignment=TA_CENTER),
        ))
    if metric_cells:
        metric_table = Table([metric_cells], colWidths=[document.width / len(metric_cells)] * len(metric_cells))
        metric_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), PALE_BLUE),
            ("BOX", (0, 0), (-1, -1), 0.5, GRID),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.extend([metric_table, Spacer(1, 7 * mm)])

    story.extend([
        Paragraph("Executive Summary", heading_style),
        Paragraph(escape(narrative.executive_summary).replace("\n", "<br/>"), body_style),
        Spacer(1, 6 * mm),
    ])

    for chart in dataset.charts:
        chart_output = render_chart(chart)
        story.append(KeepTogether([
            Image(chart_output, width=176 * mm, height=78 * mm),
            Spacer(1, 1.5 * mm),
            Paragraph("<b>What this means</b>", heading_style),
            Paragraph(escape(narrative.explanation_for(chart)), insight_style),
            Spacer(1, 5 * mm),
        ]))

    story.extend([Paragraph("Key Findings", heading_style), Spacer(1, 1 * mm)])
    for finding in narrative.key_findings:
        story.append(Paragraph(f"- {escape(finding)}", finding_style))
    story.extend([
        Spacer(1, 3 * mm),
        Paragraph("Recommended Action", heading_style),
        Paragraph(escape(narrative.recommendation), insight_style),
    ])

    document.build(story)
    return output.getvalue()


def render_xlsx_report(
    dataset: ReportDataset,
) -> bytes:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.sheet_view.showGridLines = False
    data_sheet.append(dataset.headers)
    for row in dataset.rows:
        data_sheet.append(row)
    for cell in data_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123047")
        cell.alignment = Alignment(wrap_text=True)
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions
    money_headers = {
        "Total Debt", "Amount Paid", "Balance Due", "Amount", "Service Total",
        "Logistics Total", "Total Payable", "Paid", "Available Balance",
        "Pending Balance",
    }
    for column_index, header in enumerate(dataset.headers, start=1):
        for cell in data_sheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
        ):
            for value_cell in cell:
                if header in money_headers and isinstance(value_cell.value, (int, float)):
                    value_cell.number_format = "#,##0.00"
                elif isinstance(value_cell.value, datetime):
                    value_cell.number_format = "yyyy-mm-dd hh:mm"
    for column_cells in data_sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 35)
        data_sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
