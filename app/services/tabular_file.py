import csv
import io
import re
from datetime import date, datetime

from openpyxl import load_workbook

from app.services.source_file import SourceFileError


MAX_TABULAR_ROWS = 20_000
MAX_TABULAR_CELLS = 200_000
MAX_TABULAR_TEXT_CHARACTERS = 500_000


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise SourceFileError("CSV file encoding is unsupported. Use UTF-8 or Windows-1252.")


def _cell_text(value, number_format: str = "") -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        zero_format = re.fullmatch(r"0+", number_format.strip())
        if zero_format and float(value).is_integer():
            return f"{int(value):0{len(number_format.strip())}d}"
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _append_row(lines: list[str], values: list[str], counters: dict[str, int]) -> None:
    while values and not values[-1]:
        values.pop()
    if not any(values):
        return
    counters["rows"] += 1
    counters["cells"] += len(values)
    if counters["rows"] > MAX_TABULAR_ROWS or counters["cells"] > MAX_TABULAR_CELLS:
        raise SourceFileError(
            "Spreadsheet is too large to process safely. Limit it to 20,000 rows and 200,000 cells."
        )
    lines.append("\t".join(values))


def extract_tabular_text(content: bytes, kind: str) -> str:
    lines: list[str] = []
    counters = {"rows": 0, "cells": 0}

    try:
        if kind == "csv":
            reader = csv.reader(io.StringIO(_decode_csv(content), newline=""))
            for row in reader:
                _append_row(lines, [value.strip() for value in row], counters)
        elif kind == "xlsx":
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            try:
                for worksheet in workbook.worksheets:
                    sheet_lines: list[str] = []
                    for row in worksheet.iter_rows():
                        _append_row(
                            sheet_lines,
                            [_cell_text(cell.value, cell.number_format) for cell in row],
                            counters,
                        )
                    if sheet_lines:
                        if lines:
                            lines.append("")
                        lines.append(f"--- SHEET: {worksheet.title} ---")
                        lines.extend(sheet_lines)
            finally:
                workbook.close()
        else:
            raise SourceFileError(f"Unsupported tabular source kind '{kind}'.")
    except SourceFileError:
        raise
    except Exception as exc:
        raise SourceFileError(f"Failed to read {kind.upper()} file: {str(exc)}") from exc

    text = "\n".join(lines).strip()
    if not text:
        raise SourceFileError("The uploaded CSV or XLSX file contains no readable rows.")
    if len(text) > MAX_TABULAR_TEXT_CHARACTERS:
        raise SourceFileError(
            "Spreadsheet text exceeds the 500,000 character processing limit."
        )
    return text
